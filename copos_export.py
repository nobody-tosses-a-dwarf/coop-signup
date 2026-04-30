"""
CoPOS Member Import export.

Produces a tab-delimited file matching the official 60-column CoPOS Member
Import template. Save the output as MEMBERS#.TXT and import into CoPOS.

Column legend (from template):
  Yellow (1, 23-31): New members only, ignored when updating existing members.
  Green (2-22):      Updated on existing members (matched by member number).
  Orange (32-55):    Per-payment groups (date, equity, dues) for up to six
                     actual payment records.
  Red (35, 39, 43, 47, 51, 55, 56-58): Calculated by CoPOS.
  Lt Blue (59-60):   Internal to CoPOS. Do not fill.

The signup system does not process payments, so actual-payment columns
(27-29, 32-58) are intentionally left blank. Equity Contract and Dues
Contract (cols 30-31) record the member's commitment; CoPOS tracks payments
against that contract as they come in.
"""

from datetime import datetime
from typing import List, Dict
from io import BytesIO


def _clean(value) -> str:
    """Sanitize a value for tab-delimited output.
    
    Replaces tabs and newlines with spaces to prevent breaking the row format.
    Returns '' for None. Always returns a str.
    """
    if value is None:
        return ''
    s = str(value)
    return s.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')


def _fmt_date(value) -> str:
    """Format a timestamp/date as MM/DD/YYYY. Returns '' if value is empty."""
    if not value:
        return ''
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        except ValueError:
            return value  # leave as-is if unparseable
    else:
        dt = value
    return dt.strftime('%m/%d/%Y')


def _fmt_money(value) -> str:
    """Format a numeric value as plain decimal (no dollar sign). '' if None."""
    if value is None:
        return ''
    try:
        n = float(value)
    except (TypeError, ValueError):
        return ''
    return f"{n:.2f}"


def generate_copos_export(members: List[Dict], coop: Dict) -> str:
    """
    Build a tab-delimited string matching the 60-column CoPOS template.

    One row per member. No header row is emitted (CoPOS imports start at
    row 7 in the template; the template's first six rows are documentation).
    """
    lines = []

    for m in members:
        row = [''] * 60  # 60 columns, all blank by default

        # ---- Identity (cols 1-9) ----
        row[0] = _clean(m.get('member_number'))                        # 1  Member #
        first = _clean(m.get('first_name')).strip()
        last = _clean(m.get('last_name')).strip()
        row[1] = f"{first} {last}".strip()                             # 2  Member Name #1
        row[2] = ''                                                    # 3  Member Name #2 (not collected)
        row[3] = _clean(m.get('address'))                              # 4  Street Address
        row[4] = _clean(m.get('city'))                                 # 5  City
        row[5] = _clean(m.get('state'))                                # 6  State
        row[6] = _clean(m.get('zip'))                                  # 7  Zip Code
        row[7] = _clean(m.get('phone'))                                # 8  Phone
        row[8] = _clean(m.get('email'))                                # 9  E-Mail

        # ---- Discount / privilege flags (cols 10-22) ----
        row[9]  = ''                                                   # 10 Eligible for Senior Disc?
        row[10] = ''                                                   # 11 Tax Exempt
        row[11] = 'Y' if m.get('newsletter') else 'N'                  # 12 Newsletter (from member's opt-in choice)
        row[12] = 'Y'                                                  # 13 Active
        row[13] = 'Y'                                                  # 14 Voting Privileges
        row[14] = ''                                                   # 15 Credit Limit
        row[15] = ''                                                   # 16 Special Order Disc
        row[16] = ''                                                   # 17 Basic Member Disc
        row[17] = ''                                                   # 18 Senior Discount
        row[18] = ''                                                   # 19 Working Member Discount
        row[19] = ''                                                   # 20 Working Member Discount Expires On
        row[20] = ''                                                   # 21 Employee Discount
        row[21] = ''                                                   # 22 Total Discount (calculated)

        # ---- Membership contract (cols 23-31) ----
        row[22] = _clean(m.get('membership_type_name'))                # 23 Membership Type
        row[23] = _fmt_date(m.get('signed_up_at'))                     # 24 Date Joined
        row[24] = ''                                                   # 25 Member Due Date

        plan = (m.get('payment_plan') or '').lower()
        row[25] = 'Y' if plan == 'installments' else 'N'               # 26 Paid in Installments

        equity_paid = float(m.get('equity_paid') or 0)
        pmt_date = _fmt_date(m.get('payment_date')) if equity_paid > 0 else ''

        row[26] = pmt_date                                             # 27 Initial Payment Date
        row[27] = _fmt_money(equity_paid) if plan == 'full' and equity_paid > 0 else ''        # 28 One Time Sign Up Fee Paid
        row[28] = _fmt_money(equity_paid) if plan == 'installments' and equity_paid > 0 else '' # 29 Installment Fee Paid
        row[29] = _fmt_money(m.get('total_equity'))                    # 30 Equity Contract
        row[30] = _fmt_money(m.get('total_dues', 0))                   # 31 Dues Contract

        # 1st payment slot — filled when equity was paid at signup via Stripe
        if equity_paid > 0 and pmt_date:
            row[31] = pmt_date                                         # 32 1st Payment Date
            row[32] = _fmt_money(equity_paid)                          # 33 Equity Pd In (1st)
            row[33] = '0.00'                                           # 34 Dues Pd In (1st)
        # ---- Cols 35 (Total 1st, calculated), 36-55: remaining payment slots (blank) ----
        # ---- Cols 56-58: Totals paid to date (calculated by CoPOS) ----
        # ---- Cols 59-60: CoPOS internal (must stay blank) ----

        lines.append('\t'.join(row))

    return '\n'.join(lines)


COPOS_HEADERS = [
    'Member #', 'Member Name #1', 'Member Name #2', 'Street Address', 'City',
    'State', 'Zip Code', 'Phone', 'E-Mail',
    'Eligible for Senior Disc?', 'Tax Exempt', 'Newsletter', 'Active',
    'Voting Privileges', 'Credit Limit', 'Special Order Disc', 'Basic Member Disc',
    'Senior Discount', 'Working Member Discount', 'Working Member Discount Expires On',
    'Employee Discount', 'Total Discount',
    'Membership Type', 'Date Joined', 'Member Due Date', 'Paid in Installments',
    'Initial Payment Date', 'One Time Sign Up Fee Paid', 'Installment Fee Paid',
    'Equity Contract', 'Dues Contract',
    '1st Payment Date', 'Equity Pd In (1st)', 'Dues Pd In (1st)', 'Total 1st',
    '2nd Payment Date', 'Equity Pd In (2nd)', 'Dues Pd In (2nd)', 'Total 2nd',
    '3rd Payment Date', 'Equity Pd In (3rd)', 'Dues Pd In (3rd)', 'Total 3rd',
    '4th Payment Date', 'Equity Pd In (4th)', 'Dues Pd In (4th)', 'Total 4th',
    '5th Payment Date', 'Equity Pd In (5th)', 'Dues Pd In (5th)', 'Total 5th',
    '6th Payment Date', 'Equity Pd In (6th)', 'Dues Pd In (6th)', 'Total 6th',
    'Total Equity Paid', 'Total Dues Paid', 'Total Paid',
    'CoPOS Internal 1', 'CoPOS Internal 2',
]


def generate_copos_export_xlsx(members: List[Dict], coop: Dict) -> bytes:
    """Build an Excel workbook matching the 60-column CoPOS template."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F2D', end_color='2C5F2D', fill_type='solid')

    for col_idx, header in enumerate(COPOS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Parse the tab-delimited export and write each row
    txt = generate_copos_export(members, coop)
    for row_idx, line in enumerate(txt.splitlines(), start=2):
        for col_idx, value in enumerate(line.split('\t'), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size the most useful columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
